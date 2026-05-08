"""
╔══════════════════════════════════════════════════════════════════════╗
║        Postkalkulacja Profitability  –  app.py                       ║
║        Streamlit financial dashboard for post-calculation reports    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

import io
import re
import warnings
from typing import Any

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG  (must be first Streamlit call)
# ════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Postkalkulacja Profitability",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ════════════════════════════════════════════════════════════════════════════
#  GLOBAL CSS
# ════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
/* ══════════════════════════════════════════
   LAYOUT
══════════════════════════════════════════ */
[data-testid="stAppViewContainer"]  { background: #F7EFEA; }
[data-testid="stHeader"]            { background: transparent; }
[data-testid="stToolbar"]           { display: none; }
footer                              { visibility: hidden; }
.block-container { padding-top: 1.2rem; padding-bottom: 2rem; }

/* ══════════════════════════════════════════
   SIDEBAR  — dark burgundy shell
══════════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #3A0000 0%, #5A0A0A 100%) !important;
    border-right: none !important;
    box-shadow: 4px 0 24px rgba(0,0,0,.35);
}
/* kill default radio bullets & circles */
[data-testid="stSidebar"] [data-testid="stRadio"] > div       { gap: 4px !important; }
[data-testid="stSidebar"] [data-testid="stRadio"] label > div:first-child { display: none !important; }

/* every radio label = nav button */
[data-testid="stSidebar"] [data-testid="stRadio"] label {
    display: flex !important;
    align-items: center !important;
    width: 100% !important;
    padding: 10px 14px !important;
    border-radius: 10px !important;
    font-size: 0.88rem !important;
    font-weight: 500 !important;
    color: #F5D9D0 !important;
    cursor: pointer !important;
    transition: background .15s ease, transform .1s ease !important;
    border: 1px solid transparent !important;
    margin: 1px 0 !important;
    background: transparent !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
    background: rgba(255,90,31,.22) !important;
    border-color: rgba(255,90,31,.40) !important;
    color: #FFFFFF !important;
    transform: translateX(4px) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[aria-checked="true"] {
    background: linear-gradient(90deg, #FF5A1F 0%, #C91818 100%) !important;
    border-color: transparent !important;
    color: #FFFFFF !important;
    font-weight: 700 !important;
    box-shadow: 0 3px 12px rgba(255,90,31,.45) !important;
    transform: translateX(4px) !important;
}
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,.12) !important; margin: 10px 0 !important; }
[data-testid="stSidebar"] * { color: #F5D9D0; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #FF8C66 !important; }

/* ══════════════════════════════════════════
   TYPOGRAPHY
══════════════════════════════════════════ */
h1 { color: #3A0000 !important; font-size: 1.65rem !important; font-weight: 800 !important; }
h2 { color: #6B0000 !important; font-size: 1.25rem !important; }
h3 { color: #6B0000 !important; font-size: 1.05rem !important; }

/* ══════════════════════════════════════════
   BUTTONS
══════════════════════════════════════════ */
.stButton > button {
    background: linear-gradient(135deg,#6B0000 0%,#FF5A1F 100%) !important;
    color:#fff !important; border:none !important; border-radius:9px !important;
    padding:9px 24px !important; font-weight:700 !important; font-size:.92rem !important;
    letter-spacing:.02em !important;
    transition: all .2s ease !important;
    box-shadow: 0 2px 8px rgba(107,0,0,.30) !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg,#3A0000 0%,#C91818 100%) !important;
    box-shadow: 0 5px 18px rgba(107,0,0,.50) !important;
    transform: translateY(-2px) !important;
}
.stDownloadButton > button {
    background: linear-gradient(135deg,#0FA958 0%,#0d8a47 100%) !important;
    color:#fff !important; border:none !important; border-radius:9px !important;
    padding:10px 28px !important; font-weight:700 !important; font-size:1.05rem !important;
    box-shadow: 0 2px 8px rgba(15,169,88,.35) !important;
}
.stDownloadButton > button:hover { filter: brightness(1.08); transform: translateY(-1px); }

/* ══════════════════════════════════════════
   CARDS
══════════════════════════════════════════ */
.card {
    background: #FFFFFF; border-radius: 14px;
    padding: 20px 26px; box-shadow: 0 2px 16px rgba(107,0,0,.07);
    margin-bottom: 18px; border: 1px solid rgba(107,0,0,.06);
}
.card-sm {
    background: #FFFFFF; border-radius: 10px;
    padding: 14px 18px; box-shadow: 0 1px 8px rgba(107,0,0,.06);
    margin-bottom: 12px;
}

/* ══════════════════════════════════════════
   KPI CARDS
══════════════════════════════════════════ */
.kpi {
    background: #FFF; border-radius: 14px; padding: 18px 14px;
    box-shadow: 0 3px 14px rgba(107,0,0,.10);
    border-left: 4px solid #6B0000; text-align: center;
    margin-bottom: 10px; transition: transform .15s ease;
}
.kpi:hover  { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(107,0,0,.15); }
.kpi.g  { border-left-color: #0FA958; }
.kpi.o  { border-left-color: #FF5A1F; }
.kpi.r  { border-left-color: #C91818; }
.kpi.b  { border-left-color: #1565A0; }
.kpi .v { font-size: 1.45rem; font-weight: 800; color: #3A0000; line-height: 1.1; }
.kpi .l { font-size: .68rem; color: #6B6B6B; text-transform: uppercase; letter-spacing: .07em; margin-top: 5px; }

/* ══════════════════════════════════════════
   SECTION TITLE
══════════════════════════════════════════ */
.stitle {
    color: #6B0000; font-size: 1.05rem; font-weight: 700;
    border-bottom: 2px solid #FF5A1F; padding-bottom: 5px; margin: 18px 0 13px;
}

/* ══════════════════════════════════════════
   BADGES
══════════════════════════════════════════ */
.bok   { display:inline-block; background:#0FA958; color:#fff; border-radius:20px;
         padding:2px 9px; font-size:.72rem; font-weight:700; margin-left:6px; }
.bmiss { display:inline-block; background:#C91818; color:#fff; border-radius:20px;
         padding:2px 9px; font-size:.72rem; font-weight:700; margin-left:6px; }

/* ══════════════════════════════════════════
   DATA TABLES
══════════════════════════════════════════ */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* ══════════════════════════════════════════
   TABS
══════════════════════════════════════════ */
.stTabs [data-baseweb="tab-list"] {
    background: #3A0000; border-radius: 10px 10px 0 0; gap: 2px; padding: 4px 6px;
}
.stTabs [data-baseweb="tab"] {
    color: #FAE8E0 !important; background: transparent !important;
    border-radius: 8px !important; font-weight: 500 !important; padding: 6px 14px !important;
}
.stTabs [aria-selected="true"] {
    background: #FF5A1F !important; color: #fff !important; font-weight: 700 !important;
}

/* ══════════════════════════════════════════
   MULTISELECT chips
══════════════════════════════════════════ */
[data-baseweb="tag"] {
    background: linear-gradient(135deg,#6B0000,#FF5A1F) !important;
    border-radius: 20px !important; border: none !important;
}
[data-baseweb="tag"] span { color: #fff !important; font-weight: 600 !important; }
[data-baseweb="tag"] button span { color: rgba(255,255,255,.8) !important; }

/* ══════════════════════════════════════════
   FILE UPLOADER
══════════════════════════════════════════ */
[data-testid="stFileUploader"] {
    border: 2px dashed rgba(107,0,0,.25) !important;
    border-radius: 10px !important;
    background: rgba(107,0,0,.02) !important;
    transition: border-color .2s !important;
}
[data-testid="stFileUploader"]:hover { border-color: #FF5A1F !important; }

/* ══════════════════════════════════════════
   ALERTS
══════════════════════════════════════════ */
.alert-row {
    background: #FFF3F0; border-left: 4px solid #C91818;
    border-radius: 6px; padding: 8px 14px; margin: 4px 0; font-size: .88rem;
}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ════════════════════════════════════════════════════════════════════════════

HIDDEN_COLS: set[str] = {
    "Kontrbigi [29]", "Duplex [30]", "Offset [31]",
    "Add on evey box [32]", "Autobottom / ekstra lim [33]",
    "Pantone [34]", "E-flute [35]", "Item number inlay [36]",
    "POD price [37]", "Clames [38]", "Transport [39]",
    "Kooperacja [40]", "B2 price [41]", "B1 price [42]",
    "Energia [43]", "Fix price [44]", "Cena TKW [45]",
    "Click price [46]",
    "Płyty lakierujące [21]", "Matryca Braille- Grawer [22]",
    "Patryca Braille- Wewnatrz [23]",
}

# openpyxl fill colours
FILL_HDR  = PatternFill("solid", start_color="3A0000")   # dark burgundy header
FILL_DL_H = PatternFill("solid", start_color="1565A0")   # blue DL header
FILL_DL_R = PatternFill("solid", start_color="D6ECF8")   # blue DL row tint
FILL_MT_H = PatternFill("solid", start_color="1A7A45")   # green Materials header
FILL_MT_R = PatternFill("solid", start_color="D4F0DF")   # green Materials row tint

FONT_WH  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
FONT_BD  = Font(bold=True, name="Arial", size=9)
FONT_REG = Font(name="Arial", size=9)
ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L  = Alignment(horizontal="left",   vertical="center")
ALIGN_R  = Alignment(horizontal="right",  vertical="center")
THIN_S   = Side(style="thin", color="D4D4D4")
BORDER   = Border(left=THIN_S, right=THIN_S, top=THIN_S, bottom=THIN_S)

FMT_EUR  = '#,##0.00'
FMT_PCT  = '0.0%'
FMT_DATE = 'YYYY-MM-DD'
FMT_YM   = '@'      # text
FMT_INT  = '#,##0'

# ════════════════════════════════════════════════════════════════════════════
#  SESSION STATE INITIALISATION
# ════════════════════════════════════════════════════════════════════════════

_SS_DEFAULTS: dict[str, Any] = {
    "rates":              {},
    "rates_loaded":       False,
    "click_costs": {
        "HP Indigo 7K Digital Press":  {
            "Yellow": .05, "Magenta": .05, "Cyan": .05,
            "Black": .05, "Violet": .06, "Orange": .06, "default": .05,
        },
        "HP Indigo 35K Digital Press": {
            "Yellow": .04, "Magenta": .04, "Cyan": .04,
            "Black": .04, "Violet": .05, "Orange": .05, "default": .04,
        },
    },
    "prepress":           {},
    "pp_digital":         10.0,
    "pp_offset":          40.0,
    "other_pct":          2.0,
    "tpm_thr":            60.0,
    "cm_thr":             40.0,
    "result":             None,
}
for _k, _v in _SS_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ════════════════════════════════════════════════════════════════════════════
#  PURE HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════════════════════

def norm_df(df: pd.DataFrame) -> pd.DataFrame:
    """Collapse multi-space / newline column names, strip leading/trailing."""
    df.columns = [" ".join(str(c).split()) for c in df.columns]
    return df


def fcol(df: pd.DataFrame, *candidates: str) -> str | None:
    """Return first matching column (case-insensitive, space-normalised)."""
    lc = {" ".join(c.split()).lower(): c for c in df.columns}
    for cand in candidates:
        key = " ".join(cand.split()).lower()
        if key in lc:
            return lc[key]
    return None


def sn(val: Any) -> float:
    """Safe numeric conversion; returns 0.0 on failure."""
    try:
        v = float(val)
        return 0.0 if (np.isnan(v) or np.isinf(v)) else v
    except (TypeError, ValueError):
        return 0.0


def batch_label(qty: Any) -> str:
    """Return batch range string for a quantity value."""
    try:
        q = float(str(qty).strip().replace(",", "."))
    except (TypeError, ValueError):
        return ""
    if q <= 50:        return "0-50"
    if q <= 100:       return "51-100"
    if q <= 200:       return "101-200"
    if q <= 300:       return "201-300"
    if q <= 500:       return "301-500"
    if q <= 1000:      return "501-1000"
    if q <= 1500:      return "1001-1500"
    if q <= 2000:      return "1501-2000"
    if q <= 3000:      return "2001-3000"
    if q <= 10000:     return "3001-10000"
    if q <= 20000:     return "10001-20000"
    if q <= 30000:     return "20001-30000"
    if q <= 100000:    return "30001-100000"
    return "100001-1000000"


def rate_for_machine(machine: str, rates: dict) -> float:
    """Fuzzy-match machine name to rate dict."""
    ml = machine.strip().lower()
    # 1. exact (case-insensitive)
    for k, v in rates.items():
        if k.strip().lower() == ml:
            return v
    # 2. key is substring of machine or vice versa
    for k, v in rates.items():
        kl = k.strip().lower()
        if kl in ml or ml in kl:
            return v
    # 3. word overlap (≥2 chars)
    m_words = {w for w in re.split(r"\W+", ml) if len(w) >= 2}
    best, best_v = 0, 0.0
    for k, v in rates.items():
        kw = {w for w in re.split(r"\W+", k.lower()) if len(w) >= 2}
        score = len(m_words & kw)
        if score > best:
            best, best_v = score, v
    return best_v


def classify_do(zp: str, czasy_idx: dict[str, set]) -> str:
    """
    Classify a Zlecenie produkcyjne as Digital / Offset / no printing / ''.
    czasy_idx: {zp_str -> set_of_lowercase_machine_names}
    """
    zp = str(zp).strip()
    if zp not in czasy_idx:
        return ""
    machines = czasy_idx[zp]
    HP_KW  = {"hp35k", "hp7k", "hp 35", "hp 7", "hp 1", "hp indigo"}
    HD_KW  = {"heidelberg cx 104", "heidelberg cx104", "heidelberg"}
    is_hp  = any(any(kw in m for kw in HP_KW) for m in machines)
    is_hd  = any(any(kw in m for kw in HD_KW) for m in machines)
    if is_hp:
        return "Digital"
    if is_hd:
        return "Offset"
    return "no printing"


def extract_fry_fragment(zamowienie: str) -> str:
    """
    Extract the fragment used to match against 'Nazwa linii faktury'.

    Strategy (per spec + real data analysis):
      1. Take text after '| ' in Zamówienie.
      2. Strip trailing ' [...]' bracket.
      Fallback: use 6+ char alphanumeric tokens from Zamówienie.

    Returns the best fragment to search for, or '' if none found.
    """
    s = str(zamowienie).strip()
    # Primary: after '| '
    if "| " in s:
        frag = s.split("| ", 1)[1].strip()
        # Remove trailing ' [...]'
        frag = re.sub(r"\s*\[.*$", "", frag).strip()
        if frag:
            return frag
    # Fallback: first 6+ char alphanumeric token
    tokens = re.findall(r"[A-Za-z0-9\-]{5,}", s)
    return tokens[0] if tokens else ""


# ════════════════════════════════════════════════════════════════════════════
#  FILE READERS
# ════════════════════════════════════════════════════════════════════════════

def read_post_list(uf) -> pd.DataFrame | None:
    """
    Read post_list / Baza file.
    The first 3 rows are title/meta rows; actual header is at row index 3.
    """
    if uf is None:
        return None
    try:
        raw = uf.read(); uf.seek(0)
        df = pd.read_excel(io.BytesIO(raw), header=3)
        return norm_df(df)
    except Exception as exc:
        st.error(f"❌ Błąd wczytywania Bazy: {exc}")
        return None


def read_generic(uf, sheet: int | str = 0) -> pd.DataFrame | None:
    """
    Generic reader with automatic header-row detection (skips title rows).
    """
    if uf is None:
        return None
    try:
        raw = uf.read(); uf.seek(0)
        probe = pd.read_excel(io.BytesIO(raw), sheet_name=sheet,
                              header=None, nrows=12)
        hrow = 0
        for i, row in probe.iterrows():
            vals = [str(v).strip() for v in row.dropna()]
            # A real header row has ≥3 non-null cells that don't look like
            # a report title
            if len(vals) >= 3 and not any(
                any(bad in v.lower() for bad in
                    ["export", "postkalkulacja", "lista zamkniętych",
                     "lista zamknietych", "system, z dnia"])
                for v in vals
            ):
                hrow = i
                break
        df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=hrow)
        return norm_df(df)
    except Exception as exc:
        st.warning(f"⚠️ Błąd wczytywania pliku: {exc}")
        return None


def read_stawki(uf) -> dict[str, float]:
    """Parse Stawki file → {machine_name: hourly_rate}."""
    if uf is None:
        return {}
    try:
        raw = uf.read(); uf.seek(0)
        probe = pd.read_excel(io.BytesIO(raw), header=None, nrows=12)
        hrow = 0
        for i, row in probe.iterrows():
            vals = [str(v).strip() for v in row.dropna()]
            if "Nazwa maszyny" in vals and "Stawka rbg" in vals:
                hrow = i; break
        df = pd.read_excel(io.BytesIO(raw), header=hrow)
        df = norm_df(df)
        nm = fcol(df, "Nazwa maszyny")
        sr = fcol(df, "Stawka rbg")
        if not (nm and sr):
            return {}
        return {
            str(r[nm]).strip(): sn(r[sr])
            for _, r in df.dropna(subset=[nm]).iterrows()
            if str(r[nm]).strip()
        }
    except Exception as exc:
        st.warning(f"⚠️ Błąd stawek: {exc}")
        return {}


def read_farby_pivot(uf) -> pd.DataFrame | None:
    """
    Read 'PIVOT Offset farby' (or first sheet with 'pivot'/'farb' in name)
    from the Farby file.
    """
    if uf is None:
        return None
    try:
        raw = uf.read(); uf.seek(0)
        xl = pd.ExcelFile(io.BytesIO(raw))
        sheet = next(
            (s for s in xl.sheet_names
             if "pivot" in s.lower() or "farb" in s.lower()),
            xl.sheet_names[0],
        )
        probe = pd.read_excel(io.BytesIO(raw), sheet_name=sheet,
                              header=None, nrows=8)
        hrow = 0
        for i, row in probe.iterrows():
            vals = [str(v).strip() for v in row.dropna()]
            if "Etykiety wierszy" in vals or "Suma koszt farby2" in vals:
                hrow = i; break
        df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=hrow)
        return norm_df(df)
    except Exception as exc:
        st.warning(f"⚠️ Błąd pliku Farby: {exc}")
        return None


# ════════════════════════════════════════════════════════════════════════════
#  CORE CALCULATION ENGINE
# ════════════════════════════════════════════════════════════════════════════

def build_profitability(
    uf_base, uf_czasy, uf_zlec, uf_fry, uf_inks, uf_farby,
    rates: dict, click_costs: dict, prepress: dict,
    pp_digital: float, pp_offset: float, other_pct: float,
) -> tuple[dict | None, list[str]]:
    """
    Main calculation.  Returns (result_dict, warnings_list).

    result_dict keys:
        df_prof        – main Profitability DataFrame
        df_czasy       – times DataFrame with 'Koszt pracy' column
        df_kliki       – inks DataFrame with 'Koszt klików' column
        df_farby_pivot – raw farby pivot DataFrame
        machine_cols   – list of machine-cost column names added to df_prof
        klient_col     – name of the client column (or None)
        qty_col        – name of the quantity column (or None)
        warns          – list of warning strings
    """
    warns: list[str] = []

    # ── 1. BASE ──────────────────────────────────────────────────────────────
    df = read_post_list(uf_base)
    if df is None:
        return None, ["Nie można wczytać pliku Bazy."]

    numer_c  = fcol(df, "Numer")
    zam_c    = fcol(df, "Zamówienie")
    qty_c    = fcol(df, "Zamawiana ilość", "Zamawiana ilosc")
    klient_c = fcol(df, "Klient", "Klient ID")

    for req_col in ["Numer", "Zamówienie", "Zamawiana ilość"]:
        if fcol(df, req_col) is None:
            warns.append(f"Brak wymaganej kolumny w Bazie: '{req_col}'")

    # Derived base columns
    df["Zlecenie produkcyjne"] = (
        df[numer_c].astype(str).str.split("-").str[0].str.strip()
        if numer_c else ""
    )
    df["Lewy 10"] = df[zam_c].astype(str).str[:10] if zam_c else ""
    df["Batch"]   = df[qty_c].apply(batch_label)   if qty_c else ""

    # ── 2. CZASY ─────────────────────────────────────────────────────────────
    df_czasy_raw = read_generic(uf_czasy)
    machine_cols: list[str] = []
    czasy_idx: dict[str, set] = {}   # {zp -> set of lowercase machine names}

    if df_czasy_raw is not None:
        nzp_c  = fcol(df_czasy_raw, "Numer zlecenia produkcyjnego")
        nm_c   = fcol(df_czasy_raw, "Nazwa maszyny")
        czas_c = fcol(df_czasy_raw,
                      "Czas czynnosci [min]", "Czas czynności [min]")

        if nzp_c and nm_c and czas_c:
            df_czasy_raw[nzp_c]  = df_czasy_raw[nzp_c].astype(str).str.strip()
            df_czasy_raw[nm_c]   = df_czasy_raw[nm_c].astype(str).str.strip()
            df_czasy_raw[czas_c] = pd.to_numeric(
                df_czasy_raw[czas_c], errors="coerce").fillna(0)

            # Build classification index
            for zp, grp in df_czasy_raw.groupby(nzp_c):
                czasy_idx[zp] = set(grp[nm_c].str.lower())

            # Hourly cost per row
            df_czasy_raw["_rate"] = df_czasy_raw[nm_c].apply(
                lambda m: rate_for_machine(m, rates)
            )
            df_czasy_raw["Koszt pracy"] = (
                df_czasy_raw[czas_c] / 60.0 * df_czasy_raw["_rate"]
            )

            # Pivot: one column per machine, one row per ZP
            machines = sorted(df_czasy_raw[nm_c].dropna().unique())
            machine_cols = [str(m).strip() for m in machines]

            pivot = (
                df_czasy_raw
                .pivot_table(
                    index=nzp_c, columns=nm_c,
                    values="Koszt pracy", aggfunc="sum", fill_value=0,
                )
                .reset_index()
            )
            pivot.columns = [str(c).strip() for c in pivot.columns]
            pivot.rename(columns={nzp_c: "_zp_key"}, inplace=True)

            df = df.merge(
                pivot,
                left_on="Zlecenie produkcyjne",
                right_on="_zp_key",
                how="left",
            )
            df.drop(columns=["_zp_key"], errors="ignore", inplace=True)
            for mc in machine_cols:
                if mc in df.columns:
                    df[mc] = pd.to_numeric(df[mc], errors="coerce").fillna(0)
                else:
                    df[mc] = 0.0
        else:
            warns.append("Plik Czasy: brak kolumn NZP / Nazwa maszyny / Czas.")
            df_czasy_raw = None
    elif uf_czasy:
        warns.append("Nie można wczytać pliku Czasy.")

    # Digital / Offset classification
    df["Digital/Offset"] = df["Zlecenie produkcyjne"].apply(
        lambda zp: classify_do(zp, czasy_idx)
    )

    if not machine_cols:
        warns.append(
            "Brak pliku Czasy – Total DL zostanie obliczone tylko z Prepress costs."
        )

    # ── 3. FARBY ─────────────────────────────────────────────────────────────
    df_farby_pivot = read_farby_pivot(uf_farby)

    if df_farby_pivot is not None:
        ew_c = fcol(df_farby_pivot, "Etykiety wierszy")
        kf_c = fcol(df_farby_pivot, "Suma koszt farby2")
        kp_c = fcol(df_farby_pivot, "Suma koszt płyty", "Suma koszt plyty")
        if ew_c and kf_c and kp_c:
            df_farby_pivot[ew_c] = df_farby_pivot[ew_c].astype(str).str.strip()
            df_fp = (
                df_farby_pivot[[ew_c, kf_c, kp_c]]
                .dropna(subset=[ew_c])
                .rename(columns={ew_c: "_ew",
                                  kf_c: "Offset inks",
                                  kp_c: "Płyta offsetowa"})
            )
            df = df.merge(df_fp, left_on="Lewy 10", right_on="_ew", how="left")
            df.drop(columns=["_ew"], errors="ignore", inplace=True)
        else:
            warns.append("Plik Farby: brak kolumn Etykiety / koszt farby / koszt płyty.")
    else:
        if uf_farby:
            warns.append("Nie można wczytać pliku Farby.")

    df["Offset inks"]    = pd.to_numeric(df.get("Offset inks"), errors="coerce").fillna(0)
    df["Płyta offsetowa"]= pd.to_numeric(df.get("Płyta offsetowa"), errors="coerce").fillna(0)

    # ── 4. KLIKI / INKS ──────────────────────────────────────────────────────
    df_kliki_out = None

    if uf_inks:
        try:
            raw_inks = uf_inks.read(); uf_inks.seek(0)
            df_inks  = norm_df(pd.read_excel(io.BytesIO(raw_inks)))
            jn_c  = fcol(df_inks, "Job Name")
            pn_c  = fcol(df_inks, "Press Name")
            col_c = fcol(df_inks, "Color")
            sep_c = fcol(df_inks, "Separations")

            if jn_c and sep_c:
                df_inks["Zamówienie"] = df_inks[jn_c].astype(str).str[:10]
                df_inks["_seps"] = pd.to_numeric(
                    df_inks[sep_c], errors="coerce").fillna(0)

                def _unit_cost(row) -> float:
                    mach  = str(row.get(pn_c,  "")).strip() if pn_c  else ""
                    color = str(row.get(col_c, "")).strip() if col_c else ""
                    cc = click_costs.get(mach, {})
                    return cc.get(color, cc.get("default",
                        next((v.get("default", .05) for v in click_costs.values()), .05)
                    ))

                df_inks["Koszt klików"] = (
                    df_inks.apply(_unit_cost, axis=1) * df_inks["_seps"]
                )
                grp_inks = (
                    df_inks.groupby("Zamówienie")["Koszt klików"]
                    .sum().reset_index()
                    .rename(columns={"Zamówienie": "_zk", "Koszt klików": "Moje Kliki"})
                )
                df = df.merge(grp_inks, left_on="Lewy 10", right_on="_zk", how="left")
                df.drop(columns=["_zk"], errors="ignore", inplace=True)
                df_kliki_out = df_inks.drop(columns=["_seps"], errors="ignore")
            else:
                warns.append("Plik Inks: brak kolumn Job Name / Separations.")
        except Exception as exc:
            warns.append(f"Błąd pliku Inks: {exc}")

    df["Moje Kliki"] = pd.to_numeric(df.get("Moje Kliki"), errors="coerce").fillna(0)
    kliki48_c = fcol(df, "Kliki [48]")
    if kliki48_c:
        df[kliki48_c] = pd.to_numeric(df[kliki48_c], errors="coerce").fillna(0)
        df["Kliki final"] = df[[kliki48_c, "Moje Kliki"]].max(axis=1)
    else:
        df["Kliki final"] = df["Moje Kliki"]

    # ── 5. SALES VALUE & DATA FAKTURY ────────────────────────────────────────
    df["Sales Value"]  = np.nan
    df["Data faktury"] = pd.NaT

    # ── 5a. Primary source: zlec + faktury ───────────────────────────────────
    if uf_zlec:
        df_zlec = read_generic(uf_zlec)
        if df_zlec is not None:
            nzp_z  = fcol(df_zlec, "Numer zlecenia produkcyjnego")
            wart_z = fcol(df_zlec, "Wartosc w linii FV netto",
                          "Wartość w linii FV netto")
            data_z = fcol(df_zlec, "Data wystawienia FV",
                          "Data wystawienia faktury")

            if nzp_z and wart_z:
                df_zlec["_nzp"] = df_zlec[nzp_z].astype(str).str.strip()
                agg_dict: dict = {wart_z: "sum"}
                if data_z:
                    agg_dict[data_z] = "first"
                grp_z = df_zlec.groupby("_nzp").agg(agg_dict).reset_index()

                grp_z = grp_z.rename(columns={
                    "_nzp":  "_znzp",
                    wart_z: "_sv_z",
                    **({data_z: "_dv_z"} if data_z else {}),
                })
                df = df.merge(grp_z, left_on="Zlecenie produkcyjne",
                              right_on="_znzp", how="left")

                mask_z = df["_sv_z"].notna()
                df.loc[mask_z, "Sales Value"] = df.loc[mask_z, "_sv_z"]
                if "_dv_z" in df.columns:
                    df.loc[mask_z, "Data faktury"] = pd.to_datetime(
                        df.loc[mask_z, "_dv_z"], errors="coerce"
                    )
                df.drop(
                    columns=[c for c in df.columns if c.startswith("_")],
                    errors="ignore", inplace=True,
                )

    # ── 5b. Fallback: faktury linie ───────────────────────────────────────────
    missing_mask = df["Sales Value"].isna()
    if uf_fry and missing_mask.any():
        df_fry = read_generic(uf_fry)
        if df_fry is not None:
            nl_c = fcol(df_fry, "Nazwa linii faktury")
            wl_c = fcol(df_fry, "Wartosc w linii FV netto",
                        "Wartość w linii FV netto")
            il_c = fcol(df_fry, "Ilosc w linii FV")
            dl_c = fcol(df_fry, "Data wystawienia FV")

            if nl_c and wl_c:
                fry_nl_str = df_fry[nl_c].fillna("").astype(str)
                # Convert to plain Python list for case-insensitive plain-text search
                # (avoids PyArrow regex engine errors with special chars like / [ - )
                fry_nl_lower = [s.lower() for s in fry_nl_str.tolist()]

                for idx in df[missing_mask].index:
                    zam_val = str(df.at[idx, zam_c] if zam_c else "").strip()
                    qty_val = sn(df.at[idx, qty_c] if qty_c else 0)
                    frag    = extract_fry_fragment(zam_val)

                    found_row = None
                    if frag:
                        frag_lower = frag.lower()
                        # Plain substring search — no regex, safe for all chars
                        matched_indices = [
                            i for i, s in enumerate(fry_nl_lower)
                            if frag_lower in s
                        ]
                        if matched_indices:
                            found_row = df_fry.iloc[matched_indices[0]]

                    if found_row is None:
                        continue

                    wart_fry  = sn(found_row[wl_c])
                    ilosc_fry = sn(found_row[il_c]) if il_c else 0
                    sv = (
                        (wart_fry / ilosc_fry) * qty_val
                        if ilosc_fry > 0 and qty_val > 0
                        else wart_fry
                    )
                    df.at[idx, "Sales Value"] = sv
                    if dl_c:
                        df.at[idx, "Data faktury"] = pd.to_datetime(
                            found_row[dl_c], errors="coerce"
                        )

    df["Sales Value"]  = pd.to_numeric(df["Sales Value"],  errors="coerce").fillna(0)
    df["Data faktury"] = pd.to_datetime(df["Data faktury"], errors="coerce")
    df["Miesiąc faktury"] = df["Data faktury"].dt.strftime("%Y-%m")

    # ── 6. PREPRESS COSTS ─────────────────────────────────────────────────────
    def _prepress(row) -> float:
        klient = str(row.get(klient_c, "") if klient_c else "").strip()
        oi     = sn(row.get("Offset inks", 0))
        cfg    = prepress.get(klient, {})
        if cfg:
            return cfg.get("offset" if oi > 0 else "digital",
                           pp_offset if oi > 0 else pp_digital)
        return pp_offset if oi > 0 else pp_digital

    df["Prepress costs"] = df.apply(_prepress, axis=1)

    # ── 7. OTHER MATERIALS ────────────────────────────────────────────────────
    df["Other Materials"] = df["Sales Value"] * (other_pct / 100.0)

    # ── 8. TOTAL DL ──────────────────────────────────────────────────────────
    dl_components = machine_cols + ["Prepress costs"]
    for col in dl_components:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    df["Total DL"] = df[dl_components].sum(axis=1)

    # ── 9. TOTAL MATERIALS ────────────────────────────────────────────────────
    MAT_NAMES = [
        "Papier [16]", "Klej [17]", "Lakiery [20]",
        "Opakowania zbiorcze [24]", "Other Materials",
        "Offset inks", "Płyta offsetowa", "Kliki final",
    ]
    mat_cols_used: list[str] = []
    for mc in MAT_NAMES:
        rc = fcol(df, mc)
        if rc:
            df[rc] = pd.to_numeric(df[rc], errors="coerce").fillna(0)
            mat_cols_used.append(rc)
        else:
            df[mc] = 0.0
            mat_cols_used.append(mc)
    df["Total Materials"] = df[mat_cols_used].sum(axis=1)

    # ── 10. TPM & CM ─────────────────────────────────────────────────────────
    df["TPM"] = df["Sales Value"] - df["Total DL"]
    df["CM"]  = df["Sales Value"] - df["Total DL"] - df["Total Materials"]

    # ── 11. FINAL COLUMN ORDER ────────────────────────────────────────────────
    TAIL = [
        "Total DL", "Total Materials", "Sales Value",
        "Data faktury", "Miesiąc faktury", "TPM", "CM",
    ]
    other_cols = [c for c in df.columns if c not in TAIL]
    df = df[[c for c in other_cols + TAIL if c in df.columns]]

    # Remove any stray helper columns
    df = df[[c for c in df.columns if not c.startswith("_")]]

    return {
        "df_prof":        df,
        "df_czasy":       df_czasy_raw,
        "df_kliki":       df_kliki_out,
        "df_farby_pivot": df_farby_pivot,
        "machine_cols":   machine_cols,
        "mat_cols":       mat_cols_used,
        "klient_col":     klient_c,
        "qty_col":        qty_c,
        "warns":          warns,
    }, warns


# ════════════════════════════════════════════════════════════════════════════
#  XLSX BUILDER
# ════════════════════════════════════════════════════════════════════════════

def _is_currency_col(name: str) -> bool:
    kw = [
        "wartość", "wartosc", "koszt", "sales", "tpm", "cm", "total",
        "prepress", "materials", "inks", "płyta", "kliki", "papier",
        "klej", "lak", "opak", "farby", "praca", "moje", "sprzedaż",
        "sprzedaz",
    ]
    n = name.lower()
    return any(k in n for k in kw)


def _col_group(col: str, dl_set: set, mat_set: set):
    """Return (header_fill, row_fill) for a column."""
    if col in dl_set:
        return FILL_DL_H, FILL_DL_R
    if col in mat_set:
        return FILL_MT_H, FILL_MT_R
    return FILL_HDR, None


def xlsx_write_sheet(
    ws,
    df: pd.DataFrame,
    dl_set: set = None,
    mat_set: set = None,
    hidden: set = None,
    curr_cols: set = None,
    pct_cols: set = None,
    date_cols: set = None,
    ym_cols: set = None,
    int_cols: set = None,
) -> None:
    """Write a DataFrame to an openpyxl worksheet with full formatting."""
    dl_set   = dl_set   or set()
    mat_set  = mat_set  or set()
    hidden_n = {" ".join(h.split()).lower() for h in (hidden or set())}
    cols     = list(df.columns)
    n_rows   = len(df)

    # Header row
    for ci, col in enumerate(cols, 1):
        hfill, _ = _col_group(col, dl_set, mat_set)
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill      = hfill
        cell.font      = FONT_WH
        cell.alignment = ALIGN_C
        cell.border    = BORDER

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            _, rfill = _col_group(col, dl_set, mat_set)
            val = row[col]
            if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font   = FONT_REG
            cell.border = BORDER
            if rfill:
                cell.fill = rfill
            # Number formats & alignment
            if pct_cols and col in pct_cols:
                cell.number_format = FMT_PCT
                cell.alignment = ALIGN_R
            elif date_cols and col in date_cols:
                cell.number_format = FMT_DATE
                cell.alignment = ALIGN_C
            elif ym_cols and col in ym_cols:
                cell.number_format = FMT_YM
                cell.alignment = ALIGN_C
            elif int_cols and col in int_cols:
                cell.number_format = FMT_INT
                cell.alignment = ALIGN_R
            elif curr_cols and col in curr_cols:
                cell.number_format = FMT_EUR
                cell.alignment = ALIGN_R
            else:
                cell.alignment = ALIGN_L

    # Column widths
    for ci, col in enumerate(cols, 1):
        sample = [str(df.iat[r, ci-1])[:40] for r in range(min(n_rows, 80))]
        w = max(len(str(col)), max((len(s) for s in sample), default=0)) + 2
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w, 8), 44)

    # Hide columns
    for ci, col in enumerate(cols, 1):
        if " ".join(col.split()).lower() in hidden_n:
            ws.column_dimensions[get_column_letter(ci)].hidden = True

    # Freeze + filter
    ws.freeze_panes = "A2"
    if n_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def build_xlsx(
    result: dict,
    rates: dict,
    click_costs: dict,
    prepress: dict,
    other_pct: float,
    tpm_thr: float,
    cm_thr: float,
) -> io.BytesIO:
    """Build the final XLSX workbook and return as BytesIO."""
    wb = Workbook()
    for ws in wb.worksheets:
        wb.remove(ws)

    df_prof        = result["df_prof"]
    df_czasy       = result["df_czasy"]
    df_kliki       = result["df_kliki"]
    df_farby_pivot = result["df_farby_pivot"]
    machine_cols   = result["machine_cols"]
    klient_col     = result["klient_col"]
    grp_col        = klient_col or "Zlecenie produkcyjne"

    DL_SET  = set(machine_cols + ["Prepress costs", "Total DL"])
    MAT_NAMES_STD = [
        "Papier [16]", "Klej [17]", "Lakiery [20]",
        "Opakowania zbiorcze [24]", "Other Materials",
        "Offset inks", "Płyta offsetowa", "Kliki final", "Total Materials",
    ]
    MAT_SET = {c for c in df_prof.columns
               if any(" ".join(m.split()).lower() in " ".join(c.split()).lower()
                      for m in MAT_NAMES_STD)}

    curr_c = {c for c in df_prof.columns if _is_currency_col(c)}

    # ── Profitability ─────────────────────────────────────────────────────────
    ws_p = wb.create_sheet("Profitability")
    xlsx_write_sheet(
        ws_p, df_prof,
        dl_set=DL_SET, mat_set=MAT_SET,
        hidden=HIDDEN_COLS,
        curr_cols=curr_c,
        date_cols={"Data faktury"},
        ym_cols={"Miesiąc faktury"},
    )

    # ── czasy ─────────────────────────────────────────────────────────────────
    if df_czasy is not None:
        ws_c = wb.create_sheet("czasy")
        df_c = df_czasy.drop(columns=["_rate"], errors="ignore")
        xlsx_write_sheet(ws_c, df_c,
                         curr_cols={c for c in df_c.columns if _is_currency_col(c)})

    # ── Kliki ─────────────────────────────────────────────────────────────────
    if df_kliki is not None:
        ws_k = wb.create_sheet("Kliki")
        xlsx_write_sheet(ws_k, df_kliki,
                         curr_cols={c for c in df_kliki.columns if _is_currency_col(c)})

    # ── Farby Offset ──────────────────────────────────────────────────────────
    if df_farby_pivot is not None:
        ws_f = wb.create_sheet("Farby Offset")
        xlsx_write_sheet(ws_f, df_farby_pivot,
                         curr_cols={c for c in df_farby_pivot.columns
                                    if _is_currency_col(c)})

    # ── Stawki ────────────────────────────────────────────────────────────────
    ws_st = wb.create_sheet("Stawki")
    df_st = pd.DataFrame([(k, v) for k, v in rates.items()],
                         columns=["Nazwa maszyny", "Stawka rbg (PLN/h)"])
    xlsx_write_sheet(ws_st, df_st, curr_cols={"Stawka rbg (PLN/h)"})

    # ── Koszty klików ─────────────────────────────────────────────────────────
    ws_kk = wb.create_sheet("Koszty klików")
    cc_rows = [
        {"Maszyna": k, "Kolor": c, "Koszt PLN/sep": v}
        for k, colors in click_costs.items()
        for c, v in colors.items()
    ]
    df_kk = (pd.DataFrame(cc_rows) if cc_rows
             else pd.DataFrame(columns=["Maszyna", "Kolor", "Koszt PLN/sep"]))
    xlsx_write_sheet(ws_kk, df_kk, curr_cols={"Koszt PLN/sep"})

    # ── Prepress ──────────────────────────────────────────────────────────────
    pp_rows = [{"Klient": k, "Digital": v["digital"], "Offset": v["offset"]}
               for k, v in prepress.items()]
    if not pp_rows:
        pp_rows = [{"Klient": "(domyślna)",
                    "Digital": st.session_state.get("pp_digital", 10.0),
                    "Offset":  st.session_state.get("pp_offset",  40.0)}]
    ws_pp = wb.create_sheet("Prepress")
    xlsx_write_sheet(ws_pp, pd.DataFrame(pp_rows),
                     curr_cols={"Digital", "Offset"})

    # ── Parametry ─────────────────────────────────────────────────────────────
    ws_par = wb.create_sheet("Parametry")
    df_par = pd.DataFrame([
        {"Parametr": "Other costs %",    "Wartość": other_pct / 100},
        {"Parametr": "Próg alertu TPM %","Wartość": tpm_thr   / 100},
        {"Parametr": "Próg alertu CM %", "Wartość": cm_thr    / 100},
    ])
    xlsx_write_sheet(ws_par, df_par, pct_cols={"Wartość"})

    # ── Per-month Podsumowanie + Batch ────────────────────────────────────────
    months = sorted(df_prof["Miesiąc faktury"].dropna().unique())
    for month in months:
        df_m = df_prof[df_prof["Miesiąc faktury"] == month].copy()
        if grp_col not in df_m.columns:
            continue
        df_m[grp_col] = df_m[grp_col].fillna("(brak)")

        rows_s = []
        for kl, grp in df_m.groupby(grp_col):
            sv  = grp["Sales Value"].sum()
            tpm = grp["TPM"].sum()
            cm  = grp["CM"].sum()
            rows_s.append({
                "Klient": kl, "Miesiąc": month,
                "Suma sprzedaży": sv,
                "Suma TPM":  tpm,
                "TPM %":     tpm / sv if sv else 0,
                "Suma CM":   cm,
                "CM %":      cm  / sv if sv else 0,
                "Zamówień":  len(grp),
                "Digital":   (grp["Digital/Offset"] == "Digital").sum(),
                "Offset":    (grp["Digital/Offset"] == "Offset").sum(),
                "No printing": (grp["Digital/Offset"] == "no printing").sum(),
            })
        df_s = pd.DataFrame(rows_s)
        safe_m = month.replace("/", "-")
        ws_s = wb.create_sheet(f"Podsum. {safe_m}")
        xlsx_write_sheet(ws_s, df_s,
                         curr_cols={"Suma sprzedaży","Suma TPM","Suma CM"},
                         pct_cols={"TPM %","CM %"})

        if "Batch" in df_m.columns:
            ws_b = wb.create_sheet(f"Batch {safe_m}")
            df_b = (df_m.groupby([grp_col, "Batch"])
                    .size().reset_index(name="Liczba zamówień"))
            xlsx_write_sheet(ws_b, df_b)

    # ── Kokpit ────────────────────────────────────────────────────────────────
    ws_kok = wb.create_sheet("Kokpit")
    kok_rows = []
    for month in months:
        df_m = df_prof[df_prof["Miesiąc faktury"] == month]
        sv   = df_m["Sales Value"].sum()
        tpm  = df_m["TPM"].sum()
        cm   = df_m["CM"].sum()
        kok_rows.append({
            "Miesiąc": month,
            "Sprzedaż": sv,
            "TPM": tpm,
            "TPM %": tpm / sv if sv else 0,
            "CM": cm,
            "CM %":  cm  / sv if sv else 0,
            "Klientów": df_m[grp_col].nunique() if grp_col in df_m.columns else 0,
            "Zamówień": len(df_m),
            "Digital":     (df_m["Digital/Offset"] == "Digital").sum(),
            "Offset":      (df_m["Digital/Offset"] == "Offset").sum(),
            "No printing": (df_m["Digital/Offset"] == "no printing").sum(),
        })
    if kok_rows:
        xlsx_write_sheet(ws_kok, pd.DataFrame(kok_rows),
                         curr_cols={"Sprzedaż","TPM","CM"},
                         pct_cols={"TPM %","CM %"})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════════════════════════
#  SIDEBAR NAVIGATION
# ════════════════════════════════════════════════════════════════════════════

PAGES = [
    "📂 Upload plików",
    "⚙️ Stawki maszyn",
    "🖨️ Koszty klików",
    "🎨 Prepress",
    "🔧 Parametry",
    "📋 Podgląd Profitability",
    "📈 Podsumowanie",
    "🎯 Kokpit",
    "⬇️ Pobierz XLSX",
]

# Group labels for sidebar sections
PAGE_GROUPS = {
    "⚙️ KONFIGURACJA": [0, 1, 2, 3, 4],
    "📊 ANALIZA": [5, 6, 7],
    "💾 EKSPORT": [8],
}

with st.sidebar:
    # ── Logo / title ─────────────────────────────────────────────────────
    st.markdown("""
    <div style="padding:18px 12px 8px;text-align:center;">
        <div style="font-size:2rem;margin-bottom:2px;">📊</div>
        <div style="color:#FF8C66;font-size:1.05rem;font-weight:800;letter-spacing:.04em;">
            POSTKALKULACJA
        </div>
        <div style="color:rgba(245,217,208,.55);font-size:.72rem;letter-spacing:.08em;margin-top:2px;">
            PROFITABILITY APP
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<hr style='border-color:rgba(255,255,255,.10);margin:6px 0 10px;'>", unsafe_allow_html=True)

    # ── Navigation hint ───────────────────────────────────────────────────
    st.markdown("""
    <div style="padding:0 4px 8px;color:rgba(245,217,208,.50);font-size:.68rem;
                text-transform:uppercase;letter-spacing:.1em;text-align:center;">
        ↕ wybierz zakładkę
    </div>
    """, unsafe_allow_html=True)

    # ── Grouped navigation ────────────────────────────────────────────────
    # We render grouped labels as decorative headers but use a single radio
    # for actual selection (Streamlit requires one widget).
    page = st.radio(
        "Nawigacja",
        PAGES,
        label_visibility="hidden",
        key="sidebar_nav",
    )

    st.markdown("<hr style='border-color:rgba(255,255,255,.10);margin:14px 0 10px;'>", unsafe_allow_html=True)

    # ── Status panel ──────────────────────────────────────────────────────
    base_ok   = st.session_state.get("uf_base")  is not None
    result_ok = st.session_state.get("result")   is not None

    st.markdown("""
    <div style="padding:10px 12px;background:rgba(0,0,0,.20);border-radius:10px;margin:0 4px;">
        <div style="font-size:.68rem;text-transform:uppercase;letter-spacing:.08em;
                    color:rgba(245,217,208,.50);margin-bottom:8px;">Status sesji</div>
    """, unsafe_allow_html=True)

    baza_color = "#0FA958" if base_ok else "#C91818"
    baza_icon  = "✓" if base_ok else "✗"
    baza_txt   = "wczytana" if base_ok else "brak"
    dane_color = "#0FA958" if result_ok else "#FF5A1F"
    dane_icon  = "✓" if result_ok else "⏳"
    dane_txt   = "obliczone" if result_ok else "nie obliczone"

    st.markdown(f"""
    <div style="display:flex;justify-content:space-between;align-items:center;
                padding:4px 0;border-bottom:1px solid rgba(255,255,255,.07);">
        <span style="color:rgba(245,217,208,.75);font-size:.80rem;">📋 Baza</span>
        <span style="background:{baza_color};color:#fff;border-radius:12px;
                     padding:2px 8px;font-size:.70rem;font-weight:700;">{baza_icon} {baza_txt}</span>
    </div>
    <div style="display:flex;justify-content:space-between;align-items:center;padding:4px 0;">
        <span style="color:rgba(245,217,208,.75);font-size:.80rem;">📈 Dane</span>
        <span style="background:{dane_color};color:#fff;border-radius:12px;
                     padding:2px 8px;font-size:.70rem;font-weight:700;">{dane_icon} {dane_txt}</span>
    </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        "<div style='text-align:center;color:rgba(245,217,208,.30);font-size:.65rem;"
        "letter-spacing:.05em;'>v3.1 · © Postkalkulacja</div>",
        unsafe_allow_html=True,
    )


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: UPLOAD PLIKÓW
# ════════════════════════════════════════════════════════════════════════════

if page == PAGES[0]:
    st.markdown("# 📂 Upload plików")
    st.markdown(
        '<div class="card">Uploaduj pliki źródłowe. '
        '<strong>Baza (post_list)</strong> jest wymagana — reszta jest opcjonalna. '
        'Po załadowaniu plików kliknij <strong>▶ Oblicz</strong>.</div>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 📋 Wymagane")
        st.file_uploader("Baza / post_list ⭐", type=["xlsx","xls"], key="uf_base")
        st.markdown("#### 🏭 Produkcja & Faktury")
        st.file_uploader("Czasy dla aplikacji",  type=["xlsx","xls"], key="uf_czasy")
        st.file_uploader("Zlecenia + faktury",   type=["xlsx","xls"], key="uf_zlec")
        st.file_uploader("Faktury – linie",      type=["xlsx","xls"], key="uf_fry")
    with col2:
        st.markdown("#### 💰 Koszty & Stawki")
        st.file_uploader("Kliki / Inks",               type=["xlsx","xls"], key="uf_inks")
        st.file_uploader("Stawki dla aplikacji",        type=["xlsx","xls"], key="uf_stawki")
        st.file_uploader("Farby podsumowanie (Offset)", type=["xlsx","xls"], key="uf_farby")

    # Auto-load stawki on first upload
    uf_stawki = st.session_state.get("uf_stawki")
    if uf_stawki and not st.session_state["rates_loaded"]:
        loaded = read_stawki(uf_stawki)
        if loaded:
            st.session_state["rates"] = loaded
            st.session_state["rates_loaded"] = True
            st.success(f"✅ Załadowano {len(loaded)} stawek maszyn automatycznie.")

    # Status grid
    st.markdown("---")
    st.markdown('<div class="stitle">Status plików</div>', unsafe_allow_html=True)
    FILES_STATUS = [
        ("Baza",           "uf_base"),
        ("Czasy",          "uf_czasy"),
        ("Zlecenia+FV",    "uf_zlec"),
        ("Faktury linie",  "uf_fry"),
        ("Kliki/Inks",     "uf_inks"),
        ("Stawki",         "uf_stawki"),
        ("Farby",          "uf_farby"),
    ]
    sc = st.columns(4)
    for i, (label, key) in enumerate(FILES_STATUS):
        ok = st.session_state.get(key) is not None
        badge = f'<span class="bok">✓ OK</span>' if ok else f'<span class="bmiss">✗ brak</span>'
        sc[i % 4].markdown(f"**{label}** {badge}", unsafe_allow_html=True)

    # Calculate button
    st.markdown("---")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("### ▶ Uruchom obliczenia")
    st.markdown(
        "Po kliknięciu aplikacja przetworzy wszystkie załadowane pliki "
        "i zapisze wynik w pamięci sesji."
    )
    if st.button("▶ Oblicz Profitability", use_container_width=False):
        if not st.session_state.get("uf_base"):
            st.error("❌ Brak pliku Baza! Załaduj go powyżej.")
        else:
            with st.spinner("Obliczanie…"):
                result, warns = build_profitability(
                    st.session_state.get("uf_base"),
                    st.session_state.get("uf_czasy"),
                    st.session_state.get("uf_zlec"),
                    st.session_state.get("uf_fry"),
                    st.session_state.get("uf_inks"),
                    st.session_state.get("uf_farby"),
                    st.session_state.get("rates", {}),
                    st.session_state.get("click_costs", {}),
                    st.session_state.get("prepress", {}),
                    st.session_state.get("pp_digital", 10.0),
                    st.session_state.get("pp_offset",  40.0),
                    st.session_state.get("other_pct",  2.0),
                )
            if result:
                st.session_state["result"] = result
                df = result["df_prof"]
                sv_sum = df["Sales Value"].sum()
                st.success(
                    f"✅ Obliczono {len(df):,} rekordów · "
                    f"Sprzedaż: {sv_sum:,.0f} PLN · "
                    f"Miesiące: {df['Miesiąc faktury'].nunique()}"
                )
                for w in warns:
                    st.warning(w)
            else:
                for w in warns:
                    st.error(w)
    st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: STAWKI MASZYN
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[1]:
    st.markdown("# ⚙️ Stawki maszyn")
    st.markdown(
        '<div class="card">Stawki są ładowane automatycznie z pliku '
        '<b>Stawki dla aplikacji</b>. Edytuj tabelę i kliknij Zapisz.</div>',
        unsafe_allow_html=True,
    )

    # Try to load from file if empty
    if not st.session_state["rates"] and not st.session_state["rates_loaded"]:
        uf_st = st.session_state.get("uf_stawki")
        if uf_st:
            loaded = read_stawki(uf_st)
            if loaded:
                st.session_state["rates"] = loaded
                st.session_state["rates_loaded"] = True

    rates = dict(st.session_state["rates"])

    with st.expander("➕ Dodaj maszynę"):
        c1, c2, c3 = st.columns([3, 2, 1])
        nm_new = c1.text_input("Nazwa maszyny", key="_nm_new")
        rt_new = c2.number_input("Stawka rbg (PLN/h)", value=100.0, min_value=0.0, key="_rt_new")
        if c3.button("Dodaj", key="_btn_add_mach"):
            if nm_new.strip():
                rates[nm_new.strip()] = rt_new
                st.session_state["rates"] = rates
                st.session_state["rates_loaded"] = True
                st.rerun()

    if rates:
        df_r = pd.DataFrame(
            [(k, v) for k, v in rates.items()],
            columns=["Nazwa maszyny", "Stawka rbg (PLN/h)"],
        )
        edited_r = st.data_editor(
            df_r, use_container_width=True, num_rows="dynamic", key="_rates_ed"
        )
        if st.button("💾 Zapisz stawki"):
            new_r = {}
            for _, row in edited_r.iterrows():
                nm = str(row.get("Nazwa maszyny", "")).strip()
                if nm:
                    new_r[nm] = sn(row.get("Stawka rbg (PLN/h)", 0))
            st.session_state["rates"] = new_r
            st.session_state["rates_loaded"] = True
            st.success("✅ Stawki zapisane!")
    else:
        st.info(
            "Brak stawek. Uploaduj plik 'Stawki dla aplikacji' "
            "w zakładce Upload lub dodaj ręcznie."
        )


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: KOSZTY KLIKÓW
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[2]:
    st.markdown("# 🖨️ Koszty klików")
    st.markdown(
        '<div class="card">Koszt jednostkowy (PLN) per separacja, '
        'per maszyna i kolor. HP 7K default = 0.05, HP 35K default = 0.04.</div>',
        unsafe_allow_html=True,
    )

    cc = st.session_state["click_costs"]
    cc_rows = [
        {"Maszyna (Press Name)": k, "Kolor": c, "Koszt PLN/sep": v}
        for k, colors in cc.items()
        for c, v in colors.items()
    ]
    df_cc = (
        pd.DataFrame(cc_rows) if cc_rows
        else pd.DataFrame(columns=["Maszyna (Press Name)", "Kolor", "Koszt PLN/sep"])
    )
    edited_cc = st.data_editor(
        df_cc, use_container_width=True, num_rows="dynamic", key="_cc_ed"
    )
    if st.button("💾 Zapisz koszty klików"):
        new_cc: dict = {}
        for _, row in edited_cc.iterrows():
            mach  = str(row.get("Maszyna (Press Name)", "")).strip()
            color = str(row.get("Kolor", "default")).strip()
            cost  = sn(row.get("Koszt PLN/sep", 0.05))
            if mach:
                new_cc.setdefault(mach, {})[color] = cost
        st.session_state["click_costs"] = new_cc
        st.success("✅ Koszty klików zapisane!")


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: PREPRESS
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[3]:
    st.markdown("# 🎨 Prepress")
    st.markdown(
        '<div class="card">Domyślne stawki Prepress oraz opcjonalne stawki '
        'per klient. Logika: Offset inks = 0 → Digital; Offset inks > 0 → Offset.</div>',
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    # Bind widgets directly to session_state keys (no manual assignment after widget)
    c1.number_input(
        "Domyślna stawka Digital (PLN)",
        min_value=0.0, step=1.0, key="pp_digital",
    )
    c2.number_input(
        "Domyślna stawka Offset (PLN)",
        min_value=0.0, step=1.0, key="pp_offset",
    )

    st.markdown("#### Stawki per klient (opcjonalne)")
    pp = st.session_state["prepress"]
    df_pp = (
        pd.DataFrame([{"Klient": k, "Digital": v["digital"], "Offset": v["offset"]}
                       for k, v in pp.items()])
        if pp
        else pd.DataFrame(columns=["Klient", "Digital", "Offset"])
    )
    edited_pp = st.data_editor(
        df_pp, use_container_width=True, num_rows="dynamic", key="_pp_ed"
    )
    if st.button("💾 Zapisz Prepress"):
        new_pp = {}
        for _, row in edited_pp.iterrows():
            kl = str(row.get("Klient", "")).strip()
            if kl:
                new_pp[kl] = {
                    "digital": sn(row.get("Digital", pp_dig)),
                    "offset":  sn(row.get("Offset",  pp_off)),
                }
        st.session_state["prepress"] = new_pp
        st.success("✅ Prepress zapisany!")


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: PARAMETRY
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[4]:
    st.markdown("# 🔧 Parametry")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)

    # Bind directly to session_state keys — no manual assignment after widget
    c1.number_input(
        "Other costs % (od Sales Value)",
        min_value=0.0, max_value=100.0, step=0.1, format="%.1f",
        key="other_pct",
    )
    c2.number_input(
        "Próg alertu TPM % (poniżej = alert)",
        min_value=0.0, max_value=100.0, step=1.0, format="%.0f",
        key="tpm_thr",
    )
    c3.number_input(
        "Próg alertu CM % (poniżej = alert)",
        min_value=0.0, max_value=100.0, step=1.0, format="%.0f",
        key="cm_thr",
    )

    st.markdown("""
**Objaśnienia:**
- **Other costs %** – procent od Sales Value dodawany do Total Materials jako koszty inne.
- **Próg TPM %** – jeśli TPM % klienta jest poniżej progu, klient pojawi się w tabeli alertów w Kokpicie.
- **Próg CM %** – analogicznie dla CM %.
    """)
    st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: PODGLĄD PROFITABILITY
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[5]:
    st.markdown("# 📋 Podgląd Profitability")
    result = st.session_state.get("result")

    if not result:
        st.warning("⚠️ Brak danych. Przejdź do zakładki **Upload plików** i kliknij **▶ Oblicz**.")
        st.stop()

    df_prof = result["df_prof"]
    warns   = result.get("warns", [])
    for w in warns:
        st.warning(w)

    n_rec  = len(df_prof)
    n_col  = len(df_prof.columns)
    n_mon  = df_prof["Miesiąc faktury"].nunique()
    sv_sum = df_prof["Sales Value"].sum()
    st.markdown(
        f'<div class="card"><b>{n_rec:,}</b> rekordów &nbsp;|&nbsp; '
        f'<b>{n_col}</b> kolumn &nbsp;|&nbsp; '
        f'<b>{n_mon}</b> miesięcy &nbsp;|&nbsp; '
        f'Sprzedaż: <b>{sv_sum:,.0f} PLN</b></div>',
        unsafe_allow_html=True,
    )

    # Filters
    fc1, fc2, fc3, fc4 = st.columns(4)
    months_av = sorted(df_prof["Miesiąc faktury"].dropna().unique())
    sel_m = fc1.multiselect(
        "Miesiące", ["(wszystkie)"] + months_av, default=["(wszystkie)"],
    )
    klient_c = result.get("klient_col")
    if klient_c and klient_c in df_prof.columns:
        clients_av = sorted(df_prof[klient_c].dropna().astype(str).unique())
        sel_cl = fc2.multiselect(
            "Klient", ["(wszyscy)"] + clients_av, default=["(wszyscy)"],
        )
    else:
        sel_cl = ["(wszyscy)"]

    do_av  = sorted(df_prof["Digital/Offset"].dropna().unique())
    sel_do = fc3.multiselect(
        "Digital/Offset", ["(wszystkie)"] + do_av, default=["(wszystkie)"],
    )
    batch_av  = sorted(df_prof["Batch"].dropna().unique())
    sel_bat = fc4.multiselect(
        "Batch", ["(wszystkie)"] + batch_av, default=["(wszystkie)"],
    )

    dv = df_prof.copy()
    if "(wszystkie)" not in sel_m  and sel_m:
        dv = dv[dv["Miesiąc faktury"].isin(sel_m)]
    if "(wszyscy)"   not in sel_cl and sel_cl and klient_c:
        dv = dv[dv[klient_c].astype(str).isin(sel_cl)]
    if "(wszystkie)" not in sel_do and sel_do:
        dv = dv[dv["Digital/Offset"].isin(sel_do)]
    if "(wszystkie)" not in sel_bat and sel_bat:
        dv = dv[dv["Batch"].isin(sel_bat)]

    st.markdown(f"Pokazuję **{len(dv):,}** rekordów.")
    st.dataframe(dv, use_container_width=True, height=540)


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: PODSUMOWANIE
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[6]:
    st.markdown("# 📈 Podsumowanie")
    result = st.session_state.get("result")

    if not result:
        st.warning("⚠️ Brak danych. Przejdź do zakładki **Upload plików** i kliknij **▶ Oblicz**.")
        st.stop()

    df_prof  = result["df_prof"]
    klient_c = result.get("klient_col")
    grp_col  = klient_c or "Zlecenie produkcyjne"

    months_av = sorted(df_prof["Miesiąc faktury"].dropna().unique())
    sel_months = st.multiselect(
        "Wybierz miesiące do podsumowania",
        months_av,
        default=months_av[:3] if len(months_av) >= 3 else months_av,
    )

    if not sel_months:
        st.info("Wybierz co najmniej jeden miesiąc.")
        st.stop()

    for month in sel_months:
        st.markdown(
            f'<div class="stitle">📅 {month}</div>', unsafe_allow_html=True
        )
        df_m = df_prof[df_prof["Miesiąc faktury"] == month].copy()
        if df_m.empty:
            st.info(f"Brak danych dla {month}")
            continue
        if grp_col not in df_m.columns:
            st.warning(f"Brak kolumny klienta '{grp_col}'")
            continue

        df_m[grp_col] = df_m[grp_col].fillna("(brak)").astype(str)

        rows_s = []
        for kl, grp in df_m.groupby(grp_col):
            sv  = grp["Sales Value"].sum()
            tpm = grp["TPM"].sum()
            cm  = grp["CM"].sum()
            rows_s.append({
                "Klient":           kl,
                "Sprzedaż":         round(sv, 2),
                "TPM":              round(tpm, 2),
                "TPM %":            f"{tpm/sv*100:.1f}%" if sv else "—",
                "CM":               round(cm, 2),
                "CM %":             f"{cm/sv*100:.1f}%" if sv else "—",
                "Zamówień":         len(grp),
                "Digital":          int((grp["Digital/Offset"] == "Digital").sum()),
                "Offset":           int((grp["Digital/Offset"] == "Offset").sum()),
                "No printing":      int((grp["Digital/Offset"] == "no printing").sum()),
            })
        df_s = pd.DataFrame(rows_s)

        # Totals row
        sv_t  = df_s["Sprzedaż"].sum()
        tpm_t = df_s["TPM"].sum()
        cm_t  = df_s["CM"].sum()
        df_s.loc[len(df_s)] = {
            "Klient": "━━ SUMA",
            "Sprzedaż": round(sv_t, 2),
            "TPM": round(tpm_t, 2),
            "TPM %": f"{tpm_t/sv_t*100:.1f}%" if sv_t else "—",
            "CM": round(cm_t, 2),
            "CM %": f"{cm_t/sv_t*100:.1f}%" if sv_t else "—",
            "Zamówień": int(df_s["Zamówień"].sum()),
            "Digital": int(df_s["Digital"].sum()),
            "Offset": int(df_s["Offset"].sum()),
            "No printing": int(df_s["No printing"].sum()),
        }

        sc1, sc2 = st.columns([3, 1])
        with sc1:
            st.dataframe(
                df_s.style.format({
                    "Sprzedaż": "{:,.2f}", "TPM": "{:,.2f}", "CM": "{:,.2f}",
                }),
                use_container_width=True,
            )
        with sc2:
            st.markdown("**Zamówienia · Batch × Klient**")
            if "Batch" in df_m.columns:
                df_bat = (
                    df_m.groupby([grp_col, "Batch"])
                    .size().reset_index(name="n")
                )
                st.dataframe(df_bat, use_container_width=True, height=340)


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: KOKPIT
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[7]:
    st.markdown("# 🎯 Kokpit")
    result = st.session_state.get("result")

    if not result:
        st.warning("⚠️ Brak danych. Przejdź do zakładki **Upload plików** i kliknij **▶ Oblicz**.")
        st.stop()

    df_prof  = result["df_prof"]
    klient_c = result.get("klient_col")
    grp_col  = klient_c or "Zlecenie produkcyjne"
    tpm_thr  = st.session_state["tpm_thr"]
    cm_thr   = st.session_state["cm_thr"]

    BURG, ORG, GRN, RED = "#6B0000", "#FF5A1F", "#0FA958", "#C91818"

    months_av = sorted(df_prof["Miesiąc faktury"].dropna().unique())

    # ── Filter bar ────────────────────────────────────────────────────────────
    st.markdown("""
    <div style="background:#FFFFFF;border-radius:14px;padding:16px 22px;
                box-shadow:0 2px 14px rgba(107,0,0,.08);margin-bottom:18px;
                border-left:4px solid #FF5A1F;">
        <div style="font-size:.72rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:.08em;color:#6B0000;margin-bottom:10px;">
            🔍 Filtry kokpitu
        </div>
    </div>
    """, unsafe_allow_html=True)

    frow1, frow2, frow3 = st.columns([1, 1, 2])

    with frow1:
        sel_months = st.multiselect(
            "📅 Miesiące",
            months_av,
            default=months_av[:1] if months_av else [],
            key="_kok_months",
        )

    # Build client list from selected months
    base_months = sel_months if sel_months else months_av
    df_base_filt = df_prof[df_prof["Miesiąc faktury"].isin(base_months)]
    if grp_col in df_base_filt.columns:
        all_clients = sorted(df_base_filt[grp_col].dropna().astype(str).unique())
    else:
        all_clients = []

    with frow2:
        client_mode = st.radio(
            "👤 Widok klientów",
            ["Wszyscy", "Wybrani"],
            horizontal=True,
            key="_kok_client_mode",
        )

    with frow3:
        if client_mode == "Wybrani":
            sel_clients = st.multiselect(
                f"Wybierz klientów ({len(all_clients)} dostępnych)",
                all_clients,
                default=all_clients[:3] if len(all_clients) >= 3 else all_clients,
                key="_kok_clients",
                placeholder="Kliknij i wybierz klientów…",
            )
            if not sel_clients:
                st.caption("⚠️ Nie wybrano klientów — pokazuję wszystkich.")
                sel_clients = all_clients
        else:
            sel_clients = all_clients
            n = len(all_clients)
            st.markdown(
                f'<div style="padding:8px 0;color:#6B6B6B;font-size:.85rem;">'
                f'✅ Pokazuję wszystkich <b>{n}</b> klientów</div>',
                unsafe_allow_html=True,
            )

    if not sel_months:
        st.info("Wybierz co najmniej jeden miesiąc.")
        st.stop()

    filter_clients = (client_mode == "Wybrani") and len(sel_clients) < len(all_clients)


    for month in sel_months:
        st.markdown("---")
        st.markdown(
            f'<div class="stitle">📅 {month}</div>', unsafe_allow_html=True
        )
        df_m = df_prof[df_prof["Miesiąc faktury"] == month].copy()
        if df_m.empty:
            st.info(f"Brak danych dla {month}")
            continue
        if grp_col in df_m.columns:
            df_m[grp_col] = df_m[grp_col].fillna("(brak)").astype(str)

        # Apply client filter
        if filter_clients and grp_col in df_m.columns:
            df_m = df_m[df_m[grp_col].isin(sel_clients)]
            if df_m.empty:
                st.info(f"Brak danych dla wybranych klientów w {month}")
                continue

        # Client filter info badge
        if filter_clients:
            n_shown = df_m[grp_col].nunique() if grp_col in df_m.columns else 0
            st.markdown(
                f'<div style="background:#FFF3E0;border-left:3px solid #FF5A1F;'
                f'border-radius:6px;padding:6px 14px;margin-bottom:10px;font-size:.85rem;">'
                f'🔍 Filtr aktywny: <b>{n_shown}</b> z <b>{len(all_clients)}</b> klientów</div>',
                unsafe_allow_html=True,
            )

        sv_tot  = df_m["Sales Value"].sum()
        tpm_tot = df_m["TPM"].sum()
        cm_tot  = df_m["CM"].sum()
        tpm_pct = tpm_tot / sv_tot * 100 if sv_tot else 0
        cm_pct  = cm_tot  / sv_tot * 100 if sv_tot else 0
        n_kl    = df_m[grp_col].nunique()   if grp_col in df_m.columns else 0
        n_ord   = len(df_m)
        n_dig   = int((df_m["Digital/Offset"] == "Digital").sum())
        n_off   = int((df_m["Digital/Offset"] == "Offset").sum())
        n_nop   = int((df_m["Digital/Offset"] == "no printing").sum())

        # ── KPI row 1 ────────────────────────────────────────────────────────
        k1, k2, k3, k4, k5 = st.columns(5)
        KPI1 = [
            (k1, f"{sv_tot:,.0f} PLN",  "Sprzedaż",  ""),
            (k2, f"{tpm_tot:,.0f} PLN", "TPM",        ""),
            (k3, f"{tpm_pct:.1f}%",     "TPM %",
             "g" if tpm_pct >= tpm_thr else "r"),
            (k4, f"{cm_tot:,.0f} PLN",  "CM",         ""),
            (k5, f"{cm_pct:.1f}%",      "CM %",
             "g" if cm_pct >= cm_thr else "r"),
        ]
        for col_obj, val, lbl, cls in KPI1:
            col_obj.markdown(
                f'<div class="kpi {cls}"><div class="v">{val}</div>'
                f'<div class="l">{lbl}</div></div>',
                unsafe_allow_html=True,
            )
        # ── KPI row 2 ────────────────────────────────────────────────────────
        k6, k7, k8, k9, k10 = st.columns(5)
        KPI2 = [
            (k6,  str(n_kl),  "Klientów",    "o"),
            (k7,  str(n_ord), "Zamówień",    ""),
            (k8,  str(n_dig), "Digital",     ""),
            (k9,  str(n_off), "Offset",      ""),
            (k10, str(n_nop), "No printing", ""),
        ]
        for col_obj, val, lbl, cls in KPI2:
            col_obj.markdown(
                f'<div class="kpi {cls}"><div class="v">{val}</div>'
                f'<div class="l">{lbl}</div></div>',
                unsafe_allow_html=True,
            )

        # ── Per-client aggregation ────────────────────────────────────────────
        if grp_col not in df_m.columns:
            continue

        grp = df_m.groupby(grp_col).agg(
            sv  =("Sales Value", "sum"),
            tpm =("TPM", "sum"),
            cm  =("CM", "sum"),
            n   =("Sales Value", "count"),
        ).reset_index()
        grp["tpm_pct"] = grp["tpm"] / grp["sv"].replace(0, np.nan) * 100
        grp["cm_pct"]  = grp["cm"]  / grp["sv"].replace(0, np.nan) * 100

        LAYOUT = dict(
            plot_bgcolor="white", paper_bgcolor="white",
            title_font_color=BURG, height=300,
            margin=dict(t=42, b=28, l=20, r=10),
        )

        # ── Charts row 1 ─────────────────────────────────────────────────────
        r1c1, r1c2, r1c3 = st.columns(3)
        with r1c1:
            fig = px.bar(
                grp.nlargest(5, "tpm"), x=grp_col, y="tpm",
                title="🏆 Top 5 wg TPM",
                color_discrete_sequence=[BURG],
            )
            fig.update_layout(**LAYOUT)
            st.plotly_chart(fig, use_container_width=True)
        with r1c2:
            fig = px.bar(
                grp.nlargest(5, "cm"), x=grp_col, y="cm",
                title="🏆 Top 5 wg CM",
                color_discrete_sequence=[ORG],
            )
            fig.update_layout(**LAYOUT)
            st.plotly_chart(fig, use_container_width=True)
        with r1c3:
            fig = px.bar(
                grp.nlargest(5, "sv"), x=grp_col, y="sv",
                title="🏆 Top 5 wg sprzedaży",
                color_discrete_sequence=[GRN],
            )
            fig.update_layout(**LAYOUT)
            st.plotly_chart(fig, use_container_width=True)

        # ── Charts row 2 ─────────────────────────────────────────────────────
        r2c1, r2c2, r2c3 = st.columns(3)
        with r2c1:
            do_c = df_m["Digital/Offset"].value_counts().reset_index()
            do_c.columns = ["Typ", "n"]
            fig = px.pie(
                do_c, names="Typ", values="n",
                title="🖨 Digital / Offset / No printing",
                color_discrete_sequence=[BURG, ORG, RED],
            )
            fig.update_layout(paper_bgcolor="white", title_font_color=BURG,
                              height=300, margin=dict(t=42, b=28))
            st.plotly_chart(fig, use_container_width=True)
        with r2c2:
            fig = px.bar(
                grp.sort_values("tpm_pct", ascending=False),
                x=grp_col, y="tpm_pct",
                title="TPM % wg klientów",
                color_discrete_sequence=[BURG],
            )
            fig.add_hline(y=tpm_thr, line_dash="dash", line_color=ORG,
                          annotation_text=f"Próg {tpm_thr:.0f}%",
                          annotation_font_color=ORG)
            fig.update_layout(**LAYOUT, yaxis_tickformat=".1f")
            st.plotly_chart(fig, use_container_width=True)
        with r2c3:
            fig = px.bar(
                grp.sort_values("cm_pct", ascending=False),
                x=grp_col, y="cm_pct",
                title="CM % wg klientów",
                color_discrete_sequence=[GRN],
            )
            fig.add_hline(y=cm_thr, line_dash="dash", line_color=RED,
                          annotation_text=f"Próg {cm_thr:.0f}%",
                          annotation_font_color=RED)
            fig.update_layout(**LAYOUT, yaxis_tickformat=".1f")
            st.plotly_chart(fig, use_container_width=True)

        # ── Charts row 3 ─────────────────────────────────────────────────────
        r3c1, r3c2 = st.columns(2)
        with r3c1:
            fig = px.bar(
                grp.sort_values("n", ascending=False),
                x=grp_col, y="n",
                title="Liczba zamówień wg klientów",
                color_discrete_sequence=[ORG],
            )
            fig.update_layout(**LAYOUT)
            st.plotly_chart(fig, use_container_width=True)
        with r3c2:
            if "Batch" in df_m.columns:
                df_bh = (
                    df_m.groupby([grp_col, "Batch"])
                    .size().reset_index(name="n")
                )
                pivot_hm = (
                    df_bh.pivot(index=grp_col, columns="Batch", values="n")
                    .fillna(0)
                )
                fig = px.imshow(
                    pivot_hm,
                    title="Zamówienia: Klient × Batch",
                    color_continuous_scale=["white", BURG],
                    aspect="auto",
                )
                fig.update_layout(paper_bgcolor="white", title_font_color=BURG,
                                  height=300, margin=dict(t=42, b=28))
                st.plotly_chart(fig, use_container_width=True)

        # ── Alert tables ──────────────────────────────────────────────────────
        st.markdown(
            f'<div class="stitle">⚠️ Alerty</div>', unsafe_allow_html=True
        )
        al1, al2 = st.columns(2)
        with al1:
            st.markdown(f"**TPM % poniżej {tpm_thr:.0f}%**")
            atpm = grp[grp["tpm_pct"].fillna(0) < tpm_thr].sort_values("tpm_pct")
            if not atpm.empty:
                st.dataframe(
                    atpm[[grp_col, "sv", "tpm", "tpm_pct"]].rename(columns={
                        grp_col: "Klient", "sv": "Sprzedaż",
                        "tpm": "TPM", "tpm_pct": "TPM %",
                    }).style.format({
                        "Sprzedaż": "{:,.0f}", "TPM": "{:,.0f}",
                        "TPM %": "{:.1f}%",
                    }),
                    use_container_width=True,
                )
            else:
                st.success("Brak alertów TPM ✓")
        with al2:
            st.markdown(f"**CM % poniżej {cm_thr:.0f}%**")
            acm = grp[grp["cm_pct"].fillna(0) < cm_thr].sort_values("cm_pct")
            if not acm.empty:
                st.dataframe(
                    acm[[grp_col, "sv", "cm", "cm_pct"]].rename(columns={
                        grp_col: "Klient", "sv": "Sprzedaż",
                        "cm": "CM", "cm_pct": "CM %",
                    }).style.format({
                        "Sprzedaż": "{:,.0f}", "CM": "{:,.0f}",
                        "CM %": "{:.1f}%",
                    }),
                    use_container_width=True,
                )
            else:
                st.success("Brak alertów CM ✓")


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: POBIERZ XLSX
# ════════════════════════════════════════════════════════════════════════════

elif page == PAGES[8]:
    st.markdown("# ⬇️ Pobierz XLSX")
    result = st.session_state.get("result")

    if not result:
        st.warning("⚠️ Brak danych. Przejdź do zakładki **Upload plików** i kliknij **▶ Oblicz**.")
        st.stop()

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown(
        "Kliknij **Generuj**, aby zbudować kompletny plik XLSX "
        "ze wszystkimi arkuszami i formatowaniem."
    )

    if st.button("🔄 Generuj plik XLSX"):
        with st.spinner("Budowanie pliku XLSX…"):
            xlsx_buf = build_xlsx(
                result,
                st.session_state.get("rates", {}),
                st.session_state.get("click_costs", {}),
                st.session_state.get("prepress", {}),
                st.session_state.get("other_pct", 2.0),
                st.session_state.get("tpm_thr",   60.0),
                st.session_state.get("cm_thr",    40.0),
            )
        st.success("✅ Plik gotowy!")
        st.download_button(
            label="⬇️ Pobierz Profitability.xlsx",
            data=xlsx_buf,
            file_name="Profitability.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
    st.markdown("</div>", unsafe_allow_html=True)

    # Contents table
    st.markdown("#### 📋 Zawartość pliku")
    st.markdown("""
<div class="card">
<table style="width:100%;font-size:.87rem;border-collapse:collapse;">
<tr style="background:#3A0000;color:#fff;">
  <th style="padding:8px;text-align:left">Arkusz</th>
  <th style="padding:8px;text-align:left">Opis</th>
</tr>
<tr style="background:#F7EFEA">
  <td style="padding:7px"><b>Profitability</b></td>
  <td style="padding:7px">Główny arkusz · nagłówki kolorowane (niebieski = DL, zielony = Materials)
  · ukryte kolumny [29]–[46] · filtry · zamrożony wiersz 1</td>
</tr>
<tr>
  <td style="padding:7px"><b>czasy</b></td>
  <td style="padding:7px">Surowe dane z pliku czasów + wyliczony Koszt pracy (PLN) per wiersz</td>
</tr>
<tr style="background:#F7EFEA">
  <td style="padding:7px"><b>Kliki</b></td>
  <td style="padding:7px">Surowe dane Inks + wyliczony Koszt klików per wiersz</td>
</tr>
<tr>
  <td style="padding:7px"><b>Farby Offset</b></td>
  <td style="padding:7px">Pivot farby z pliku Farby podsumowanie</td>
</tr>
<tr style="background:#F7EFEA">
  <td style="padding:7px"><b>Stawki</b></td>
  <td style="padding:7px">Stawki rbg maszyn (PLN/h)</td>
</tr>
<tr>
  <td style="padding:7px"><b>Koszty klików</b></td>
  <td style="padding:7px">Koszty per separacja · per maszyna i kolor</td>
</tr>
<tr style="background:#F7EFEA">
  <td style="padding:7px"><b>Prepress</b></td>
  <td style="padding:7px">Stawki Prepress Digital / Offset per klient</td>
</tr>
<tr>
  <td style="padding:7px"><b>Parametry</b></td>
  <td style="padding:7px">Other costs %, progi alertów TPM i CM</td>
</tr>
<tr style="background:#F7EFEA">
  <td style="padding:7px"><b>Podsum. YYYY-MM</b></td>
  <td style="padding:7px">Osobny arkusz podsumowania per miesiąc · Klient · TPM% · CM%</td>
</tr>
<tr>
  <td style="padding:7px"><b>Batch YYYY-MM</b></td>
  <td style="padding:7px">Zestawienie Klient × Batch per miesiąc</td>
</tr>
<tr style="background:#F7EFEA">
  <td style="padding:7px"><b>Kokpit</b></td>
  <td style="padding:7px">Zbiorczy KPI dla wszystkich miesięcy</td>
</tr>
</table>
</div>
""", unsafe_allow_html=True)

    # Quick stats
    df_prof = result["df_prof"]
    st.markdown("#### 📊 Szybkie statystyki")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Rekordów",    f"{len(df_prof):,}")
    m2.metric("Miesięcy",     df_prof["Miesiąc faktury"].nunique())
    m3.metric("Sprzedaż (PLN)", f"{df_prof['Sales Value'].sum():,.0f}")
    m4.metric("TPM (PLN)",    f"{df_prof['TPM'].sum():,.0f}")
